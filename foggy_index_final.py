
__author__ = '牟云瀚&杜鸿宇'
import openpyxl
import string
import jieba #导入jieba分词包
import time
time_start=time.time()

from openpyxl import Workbook
from openpyxl import load_workbook
#录入各种路径
data_file_name = r'/Users/dhy.scu/Desktop/招股说明书.xlsx' #录入段落所在excel
data_sheet_name = 'Sheet1' #定义数据所在页
output_file_name = '/Users/dhy.scu/Desktop/result_output.xlsx' #录入计算结果输出的excel
output_sheet_name = 'Sheet1' #定义数据所在页
totallines = 101
#录入分词词库
file_pro_word_cut_list = "/Users/dhy.scu/Desktop/经济 财政 金融 证券 货币 商品 市场 外汇_总.txt"
#录入判别用词库
file_pro_word_judge_list = r'/Users/dhy.scu/Desktop/经济 财政 金融 证券 货币 商品 市场 外汇_分.txt'
file_difficult_char_list = r'/Users/dhy.scu/Desktop/difficult_char.txt'
file_sym_char_list = r'/Users/dhy.scu/Desktop/sym_char.txt'
file_logic_list = r'/Users/dhy.scu/Desktop/logic.txt'
file_chengyu_list = r'/Users/dhy.scu/Desktop/成语.txt'
#录入storkes的路径
strokes_path = '/Users/dhy.scu/Desktop/Strokes.txt'

beginning_1 = ['1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','一','二','三','四','五','A','B','C','D']
#文本开头需要删除的数字和字母
beginning_2 = ['*','‘','’',',','。','；','：','\n','《》','【】','[]','{}','｛｝','()','（）',' ','，','“','”']

#去除标点的函数
def strclear(text,newsigh=''):
	signtext = string.punctuation + newsigh
	signrepl = '@' * len(signtext)
	signtable = str.maketrans(signtext, signrepl)
	return text.translate(signtable).replace('@', '')
#对比库并求比例的函数
def word_ratio(segment_list,judge_package):
	with open(judge_package, 'r',encoding='utf-8') as rf:
		str = rf.read()
		#print(str.splitlines())
	x_seg_list_count = len(segment_list)
	x_corres_word_count = 0
	for j in range(x_seg_list_count):
		if segment_list[j] in str.splitlines():
			x_corres_word_count = x_corres_word_count + 1
	x_corres_word_count= x_corres_word_count / x_seg_list_count
	return x_corres_word_count
#获取笔画数目的函数
def get_stroke(c):
    # 如果返回 0, 则也是在unicode中不存在kTotalStrokes字段
    strokes = []
    #录入Strokes.txt对应的路径，注意路径中不能存在中文
    #strokes_path = 'C:\\Users\\Jackie\\Desktop\\Strokes.txt'
    with open(strokes_path, 'r') as fr:
        for line in fr:
            strokes.append(int(line.strip()))

    unicode_ = ord(c)

    if 13312 <= unicode_ <= 64045:
        return strokes[unicode_-13312]
    elif 131072 <= unicode_ <= 194998:
        return strokes[unicode_-80338]
    else:
        #print("c should be a CJK char, or not have stroke in unihan data.")
        # can also return 0
        return 0
data_excel_file = load_workbook(data_file_name) #打开所在页
data_sheet1 = data_excel_file[data_sheet_name] #再定义所在页

output_file = load_workbook(output_file_name) #打开输出文件
output_file_sheet = output_file[output_sheet_name]  #打开输出文件对应sheet

jieba.load_userdict(file_pro_word_cut_list) #采用规定的词库

for i in range(100,totallines):
    x = data_sheet1.cell(i,6).value #具体单元格
    #x_clear_1 = strclear(x , '*‘’\'"“”，,。；：\nn《》【】[]{}｛｝()（） ')
    x_clear = strclear(x , '\n ')

    #段落分成单个字并存成list
    x_char_list = []
    for k in range(0,len(x_clear)):
        x_char_list.append(x_clear[k])
    for j in range(len(x_char_list),-1,1):
            if x_char_list[j] in beginning_2:
                    x_char_list.remove(x_char_list[j])
    for j in range(0,2):
            if x_char_list[j] in beginning_1:
                    x_char_list.remove(x_char_list[j])
    #print(x_char_list)

    #段落分成词汇并存成list（基于jieba分词包及自定义的追加字库）
    #jieba.load_userdict(file_pro_word_cut_list) #采用规定的词库
    # x_seg_generator = jieba.cut(x_clear, cut_all=False) #jieba分词
    # print('Default Mode:' + '\n'.join(x_seg_generator)) #给出分词后的结果
    x_seg_list = jieba.lcut(x_clear) #写成列表的形式
    for j in range(len(x_seg_list),-1,1):
            if x_seg_list[j] in beginning_2:
                    x_seg_list.remove(x_seg_list[j])
    for j in range(0,2):
            if x_seg_list[j] in beginning_1:
                    x_seg_list.remove(x_seg_list[j])
    #print(x_seg_list)

    #计算自变量的取值
    #专业词汇比例
    x_pro_word_ratio = word_ratio(x_seg_list,file_pro_word_judge_list)
    #print(x_pro_word_ratio)
    output_file_sheet.cell(i,1,value = x_pro_word_ratio)

    #难字比例
    x_difficult_char_ratio = word_ratio(x_char_list,file_difficult_char_list)
    #print(x_difficult_char_ratio)
    output_file_sheet.cell(i, 2, value=x_difficult_char_ratio)

    #对称字
    x_sym_char_ratio = word_ratio(x_char_list, file_sym_char_list)
    #print(x_sym_char_ratio)
    output_file_sheet.cell(i, 3, value=x_sym_char_ratio)

    #逻辑词
    x_logic_ratio = word_ratio(x_seg_list, file_logic_list)
    #print(x_logic_ratio)
    output_file_sheet.cell(i, 4, value=x_logic_ratio)

    #成语比例
    x_chengyu_ratio = word_ratio(x_seg_list, file_chengyu_list)
    output_file_sheet.cell(i, 5, value=x_chengyu_ratio)

    #笔画相关
    #if pp == 1:
    x_bihua_data = x_char_list
    bihua_05 = 0
    bihua_510 = 0
    bihua_1015 = 0
    bihua_15up = 0
    bihua_total = 0
    length_data = len(x_bihua_data)
    for j in range(0, length_data):
        bihua_j = get_stroke(x_bihua_data[j])
        bihua_total = bihua_total + bihua_j
        if (bihua_j < 6):
            bihua_05 = bihua_05 + 1
        elif (bihua_j < 11):
            bihua_510 = bihua_510 + 1
        elif (bihua_j < 16):
            bihua_1015 = bihua_1015 + 1
        else:
            bihua_15up = bihua_15up + 1
    bihua_average = bihua_total / length_data  # 平均笔画数
    bihua_05_ratio = bihua_05 / length_data  # 0-5画比例
    bihua_510_ratio = bihua_510 / length_data  # 6-10画比例
    bihua_1015_ratio = bihua_1015 / length_data  # 11-15画比例
    bihua_15up_ratio = bihua_15up / length_data  # 16+画比例
    output_file_sheet.cell(i, 7, value=bihua_average)
    output_file_sheet.cell(i, 8, value=bihua_05_ratio)
    output_file_sheet.cell(i, 9, value=bihua_510_ratio)
    output_file_sheet.cell(i, 10, value=bihua_1015_ratio)
    output_file_sheet.cell(i, 11, value=bihua_15up_ratio)
    # print(bihua_average)
    # print(bihua_05_ratio)
    # print(bihua_510_ratio)
    # print(bihua_1015_ratio)
    # print(bihua_15up_ratio)

    #单字/双字/多字词
    x_single_char_word = 0
    x_double_char_word = 0
    x_mulit_char_word = 0
    x_seg_list_count = len(x_seg_list)
    for j in range(0, x_seg_list_count):
        word_len = len(x_seg_list[j])
        if word_len == 1:
            x_single_char_word = x_single_char_word + 1
        elif word_len == 2:
            x_double_char_word = x_double_char_word + 1
        else:
            x_mulit_char_word = x_mulit_char_word + 1

    x_single_char_word_ratio = x_single_char_word / x_seg_list_count
    x_double_char_word_ratio = x_double_char_word / x_seg_list_count
    x_mulit_char_word_ratio = x_mulit_char_word / x_seg_list_count
    output_file_sheet.cell(i, 12, value=x_single_char_word_ratio)
    output_file_sheet.cell(i, 13, value=x_double_char_word_ratio)
    output_file_sheet.cell(i, 14, value=x_mulit_char_word_ratio)
    # print(x_single_char_word_ratio)
    # print(x_double_char_word_ratio)
    # print(x_mulit_char_word_ratio)
    print(i)
output_file.save(output_file_name)
time_end = time.time()
print('totally cost',time_end-time_start)
print('程序制作人：' + __author__)
exit('计算完成')
